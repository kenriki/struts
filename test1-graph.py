import openpyxl


# データとなるExcelファイルを開きます
workbook = openpyxl.load_workbook('./data.xlsx')
sheet = workbook["Sheet1"]

# #####################################

# 請求月 

# #####################################

month = []
 
for i in range(3,17):
    cell_value = sheet.cell(row=i, column=2).value
 
    if cell_value not in month:
        month.append(cell_value)
 
print(month)

suppliers = month

# #####################################

# 電気代 

# #####################################

ene_price = []
 
for i in range(3,17):
    cell_value = sheet.cell(row=i, column=6).value
 
    if cell_value not in ene_price:
        ene_price.append(cell_value)
 
print(ene_price)

# #####################################

# ガス代 

# #####################################

gas_price = []
 
for i in range(3,17):
    cell_value = sheet.cell(row=i, column=7).value
 
    if cell_value not in gas_price:
        gas_price.append(cell_value)
 
print(gas_price)

# #####################################

# 水道代 

# #####################################
wat_price = []
 
for i in range(3,17):
    cell_value = sheet.cell(row=i, column=8).value
 
    if cell_value not in wat_price:
        wat_price.append(cell_value)
 
print(wat_price)


# #####################################

# 合計 

# #####################################

transaction_amounts = []

min_row = 0
max_row = 0
min_row = sheet.min_row
max_row = sheet.max_row
 
for i in suppliers:
    transaction_amount = 0
    
    for i in range(min_row, max_row+1):
        transaction_amount = "=SUM(B{}:C{}:D{})".format(i,i,i)
            
    transaction_amounts.append(transaction_amount)
 
print(transaction_amounts)
 
new_sheet = workbook.create_sheet("集計")
 
print(workbook.worksheets)
 
new_sheet["A1"] = "請求月"
new_sheet["B1"] = "電気代"
new_sheet["C1"] = "ガス代"
new_sheet["D1"] = "水道代"
new_sheet["E1"] = "総支払額"
 
for i, j, k,l,m,n in zip(list(range(2,17)), 
    month, ene_price,gas_price,wat_price,transaction_amounts):
    
    new_sheet.cell(row=i, column=1, value=j)
    new_sheet.cell(row=i, column=2, value=k)
    new_sheet.cell(row=i, column=3, value=l)
    new_sheet.cell(row=i, column=4, value=m)
    new_sheet.cell(row=i, column=5, value=n)
 
print(list(new_sheet.values))

# #####################################

# グラフ作成 

# #####################################

from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border

sheetNames = workbook.sheetnames

#del workbook['Sheet1']
#ws = workbook["集計"] 

ws = workbook['Sheet1']

# 出力するグラフのサイズ
chtWidth = 24
chtHeight = 12

# 1つ目の散布図を用意します
cht1 = openpyxl.chart.ScatterChart()
cht1.y_axis.axId = 200
cht1.y_axis.title = '上昇率'
cht1.height = chtHeight
cht1.width = chtWidth

# 2つ目の散布図を用意します
cht2 = openpyxl.chart.ScatterChart()
cht2.title = '年間光熱費'
cht2.x_axis.title = '請求月'
cht2.y_axis.title = '支払額'
cht2.y_axis.majorGridLines = None
cht2.height = chtHeight
cht2.width = chtWidth

# グラフ化するデータを参照する
startRow = 1
endRow = 60

# Excelのデータ参照した変数を用意する
frameValues = openpyxl.chart.Reference(ws, min_col=2, min_row=startRow, max_row=endRow)
heightValues = openpyxl.chart.Reference(ws, min_col=5, min_row=startRow, max_row=endRow)
wPosXValues = openpyxl.chart.Reference(ws, min_col=7, min_row=startRow, max_row=endRow)
wPosYValues = openpyxl.chart.Reference(ws, min_col=8, min_row=startRow, max_row=endRow)
moveDistValues = openpyxl.chart.Reference(ws, min_col=6, min_row=startRow, max_row=endRow)

title1=sheet.cell(row=1, column=1).value
title2=sheet.cell(row=1, column=2).value
title3=sheet.cell(row=1, column=3).value
title4=sheet.cell(row=1, column=4).value

# 系列を用意し、データ参照を入力する 最初の行を凡例とする
s1 = openpyxl.chart.Series(heightValues, frameValues, title_from_data=False, title=title1)
s2 = openpyxl.chart.Series(wPosXValues, frameValues, title_from_data=False, title=title2)
s3 = openpyxl.chart.Series(wPosYValues, frameValues, title_from_data=False, title=title3)
s4 = openpyxl.chart.Series(moveDistValues, frameValues, title_from_data=False, title=title4)

# グラフの書式設定をする
s1.graphicalProperties.line.solidFill = "4f81bd" # グラフの線の色
s1.marker.symbol = "diamond" # 各データ点のマーカーの形状
s1.marker.graphicalProperties.solidFill = "4f81bd" # 各データ点のマーカーの塗りつぶし色
s1.marker.graphicalProperties.line.solidFill = "4f81bd" # 各データ点のマーカーの枠の色

s2.graphicalProperties.line.solidFill = "8064a2"
s2.marker.symbol = "triangle"
s2.marker.graphicalProperties.solidFill = "8064a2"
s2.marker.graphicalProperties.line.solidFill = "8064a2"

s3.graphicalProperties.line.solidFill = "9bbb59"
s3.marker.symbol = "triangle"
s3.marker.graphicalProperties.solidFill = "9bbb59"
s3.marker.graphicalProperties.line.solidFill = "9bbb59"

s4.graphicalProperties.line.solidFill = "c0504d"
s4.marker.symbol = "x"
s4.marker.graphicalProperties.solidFill = "c0504d"
s4.marker.graphicalProperties.line.solidFill = "c0504d"

# Chartに系列を追加する
# 1つ目のグラフに系列(s1)を、2つ目のグラフに3つの系列(s2, s3, s4)を追加している。
cht1.series.append(s1)
cht2.series.append(s2)
cht2.series.append(s3)
cht2.series.append(s4)

# y軸を2軸もつグラフに設定する(グラフを足し合わせる)
# 2つ目のグラフのy軸を右側に設定する
cht2.y_axis.crosses = "max"
cht2 += cht1

# Excelシートにグラフを貼り付ける
graphInsertCol = 1 # グラフを挿入する列番号
inColLetter = get_column_letter(graphInsertCol)
inRow = 20 # グラフを挿入する行番号
inCellLetter = inColLetter + str(inRow) # グラフを挿入するセルの位置をExcel形式で作る
ws.add_chart(cht2, inCellLetter)

# #####################################

# 出力したいExcelファイル名を指定 

# #####################################

workbook.save('./result.xlsx')